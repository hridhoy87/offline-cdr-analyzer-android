"""
Offline Intelligence Engine for CDR Analysis.
"""
import os
import json
import string
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
import time
from difflib import SequenceMatcher
import pdf_converter  # Import the new PDF converter

# Load TAC Database for hardware insights
TAC_DB = None
def lookup_imei(imei):
    global TAC_DB
    if not imei or len(str(imei)) < 8: return None
    try:
        if TAC_DB is None:
            db_path = os.path.join(os.path.dirname(__file__), "tac_db.csv")
            if os.path.exists(db_path):
                # Using standard open
                with open(db_path, 'r', encoding='utf-8') as f:
                    TAC_DB = pd.read_csv(f, dtype=str)
                TAC_DB['TAC'] = TAC_DB['TAC'].str.strip()
            else: return None
        
        tac = str(imei)[:8]
        match = TAC_DB[TAC_DB['TAC'] == tac]
        if not match.empty:
            row = match.iloc[0]
            return f"{row['Manufacturer']} {row['Model']}"
    except: pass
    return None

def load_aliases():
    try:
        possible_paths = [
            "/storage/emulated/0/Documents/CDR_Reports/aliases_metadata.json",
            os.path.join(os.environ.get("HOME", "/storage/emulated/0"), "Documents/CDR_Reports/aliases_metadata.json"),
            os.path.join(os.path.dirname(__file__), "aliases_metadata.json")
        ]
        for p in possible_paths:
            if os.path.exists(p):
                with open(p, 'r', encoding='utf-8') as f:
                    return json.load(f)
    except: pass
    return {}

def format_with_alias(val, alias_map):
    s_val = str(val).strip()
    if s_val in alias_map:
        return f"{s_val}(📌 {alias_map[s_val]})"
    return s_val

def process_cdr_data(file_paths, intended_location, output_dir, start_ts=None, end_ts=None):
    try:
        if not file_paths: return {"status": "error", "message": "No files selected."}
        alias_map = load_aliases()
        all_dataframes, number_source_map = [], {}
        is_single_file = (len(list(file_paths)) == 1)

        def clean_phone_number(val):
            val_str = str(val).strip().split('.')[0]
            if val_str.startswith('+88'): val_str = val_str[3:]
            elif val_str.startswith('88'): val_str = val_str[2:]
            digits_only = "".join(c for c in val_str if c.isdigit())
            if len(digits_only) == 11: return digits_only
            if val_str in ['nan', 'None', '', 'nan.0']: return "[?] Unknown"
            return f"[?] {val_str}"

        for path in file_paths:
            if not os.path.exists(path): continue
            filename = os.path.basename(path)
            
            # Handle PDF Files
            if path.lower().endswith('.pdf'):
                try:
                    df = pdf_converter.pdf_to_dataframe(path)
                    if not df.empty:
                        # Extract B-Parties for the number_source_map (same as excel logic)
                        # In the standard 13-column format, Col D is index 3
                        col_D_name = df.columns[3]
                        def get_only_valid(v):
                            c = "".join(i for i in str(v).split('.')[0] if i.isdigit())
                            if c.startswith('88'): c = c[2:]
                            if c.startswith('880'): c = c[3:]
                            return c if len(c) == 11 else None
                        
                        df_clean_temp = df[col_D_name].apply(get_only_valid)
                        for num in df_clean_temp.dropna().unique():
                            if num not in number_source_map: number_source_map[num] = set()
                            number_source_map[num].add(filename)
                        all_dataframes.append(df)
                except Exception as e:
                    print(f"Error processing PDF {filename}: {e}")
                continue

            # Handle Excel Files
            try:
                excel_file = pd.ExcelFile(path, engine="openpyxl")
                df = excel_file.parse(sheet_name=excel_file.sheet_names[0], dtype=str)
                if not df.empty and len(df.columns) >= 4:
                    col_D_name = df.columns[3]
                    def get_only_valid(v):
                        c = "".join(i for i in str(v).split('.')[0] if i.isdigit())
                        if c.startswith('88'): c = c[2:]
                        if c.startswith('880'): c = c[3:]
                        return c if len(c) == 11 else None
                    df_clean_temp = df[col_D_name].apply(get_only_valid)
                    for num in df_clean_temp.dropna().unique():
                        if num not in number_source_map: number_source_map[num] = set()
                        number_source_map[num].add(filename)
                    all_dataframes.append(df)
            except: continue

        if not all_dataframes: return {"status": "error", "message": "No readable data found."}
        combined_df = pd.concat(all_dataframes, ignore_index=True).dropna(how='all')
        if len(combined_df.columns) < 13: return {"status": "error", "message": "Invalid CDR format."}

        col_A, col_C, col_D, col_E, col_F, col_H, col_I, col_J, col_L = combined_df.columns[0], combined_df.columns[2], combined_df.columns[3], combined_df.columns[4], combined_df.columns[5], combined_df.columns[7], combined_df.columns[8], combined_df.columns[9], combined_df.columns[11]

        if start_ts and end_ts:
            temp_time = pd.to_datetime(combined_df[col_A], errors='coerce')
            mask = (temp_time >= pd.to_datetime(start_ts)) & (temp_time <= pd.to_datetime(end_ts))
            combined_df = combined_df[mask]
            if combined_df.empty: return {"status": "error", "message": "No data in timeline."}

        combined_df[col_C] = combined_df[col_C].apply(clean_phone_number)
        combined_df[col_D] = combined_df[col_D].apply(clean_phone_number)
        for col in [col_F, col_H, col_I, col_L, col_J]: combined_df[col] = combined_df[col].fillna('').astype(str).str.strip().replace('nan', '')
        for col in [col_H, col_I, col_J]: combined_df[col] = combined_df[col].apply(lambda x: str(x).split('.')[0])
        
        analysis_df = combined_df[ (~combined_df[col_C].str.startswith('[?]', na=False)) & (~combined_df[col_D].str.startswith('[?]', na=False)) ]
        b_to_as = analysis_df.groupby(col_D)[col_C].nunique()
        extracted_common = b_to_as[b_to_as > 1].index.tolist()

        combined_df['_parsed_dt'] = pd.to_datetime(combined_df[col_A], errors='coerce')
        unique_a_parties = [num for num in combined_df[col_C].unique() if num and str(num) != 'nan']
        summary_a_parties_str = ", ".join([format_with_alias(n, alias_map) for n in unique_a_parties])

        raw_imei_clean = combined_df[(combined_df[col_J] != '') & (combined_df[col_C] != '')]
        target_imei_counts = raw_imei_clean.groupby(col_C)[col_J].nunique().to_dict()
        true_swapped = [f"{format_with_alias(n, alias_map)} ({c} profiles)" for n, c in target_imei_counts.items() if c >= 3]
        summary_imei_swappers_str = f"IMEI Swappers: {', '.join(true_swapped)}" if true_swapped else "Hardware Stability: No device swapping observed."

        imei_sim_mapping = raw_imei_clean.groupby(col_J)[col_C].nunique().to_dict()
        true_multi = [f"Handset {format_with_alias(i, alias_map)} ({s} numbers)" for i, s in imei_sim_mapping.items() if s >= 3]
        summary_multi_sim_str = f"Multi-SIM Burners: {', '.join(true_multi)}" if true_multi else "Device Identity: No multi-SIM handset anomalies."

        imei_to_aps = raw_imei_clean.groupby(col_J)[col_C].unique().apply(list).to_dict()
        detailed_imei_map = {str(imei): {"sims": [str(a) for a in aps], "hardware": lookup_imei(imei) or "Unknown Device", "alias": alias_map.get(str(imei), "")} for imei, aps in imei_to_aps.items() if len(aps) > 1}

        ap_to_imeis = raw_imei_clean.groupby(col_C)[col_J].unique().apply(list).to_dict()
        sim_to_imei_data = {"links": [], "nodes": []}
        seen_nodes = set()
        for ap, imeis in ap_to_imeis.items():
            if len(imeis) > 1:
                sap = str(ap)
                if sap not in seen_nodes:
                    sim_to_imei_data["nodes"].append({"id": sap, "type": "SIM", "alias": alias_map.get(sap, "")})
                    seen_nodes.add(sap)
                for imei in imeis:
                    simei = str(imei)
                    if simei not in seen_nodes:
                        sim_to_imei_data["nodes"].append({"id": simei, "type": "IMEI", "alias": alias_map.get(simei, "")})
                        seen_nodes.add(simei)
                    sim_to_imei_data["links"].append({"source": sap, "target": simei})

        night_stays_list, deep_night_count, total_night_count, raw_night_indices = [], 0, 0, []
        for idx, ts in enumerate(combined_df['_parsed_dt']):
            if pd.notnull(ts):
                if ts.hour >= 18 or ts.hour < 6: 
                    raw_night_indices.append(idx); total_night_count += 1
                if 1 <= ts.hour <= 4: deep_night_count += 1

        if raw_night_indices:
            valid_locs = combined_df.iloc[raw_night_indices][col_L].astype(str).str.strip()
            valid_locs = valid_locs[valid_locs != '']
            if not valid_locs.empty:
                for addr_text, count in valid_locs.value_counts().head(5).items():
                    matching = combined_df.iloc[raw_night_indices][combined_df.iloc[raw_night_indices][col_L] == addr_text]
                    tower = matching.groupby([col_H, col_I]).size().reset_index(name='count').sort_values(by='count', ascending=False).iloc[0]
                    night_stays_list.append(f"{addr_text} [LAC: {tower[col_H]}, Cell: {tower[col_I]}]")

        summary_night_stays_str = " | ".join(night_stays_list) if night_stays_list else "Insufficient Data"
        deep_night_pct = round((deep_night_count / total_night_count) * 100, 1) if total_night_count > 0 else 0
        summary_night_routine_str = f"Critical Windows: {deep_night_pct}% of night actions occurred between 01:00 AM and 04:00 AM."

        final_condition = pd.Series(True, index=combined_df.index)
        if intended_location:
            keywords = [loc.strip().lower() for loc in intended_location.split(",") if loc.strip()]
            if keywords: final_condition = combined_df[col_L].str.lower().apply(lambda val: any(k in val for k in keywords))

        filtered_df = combined_df[final_condition].copy()
        if filtered_df.empty: return {"status": "error", "message": "No rows match."}
        
        valid_b_mask = ~filtered_df[col_D].str.startswith('[?]', na=False)
        filtered_df[col_E] = pd.to_numeric(filtered_df[col_E], errors="coerce").fillna(0)
        freq_map = filtered_df[valid_b_mask][col_D].value_counts()
        dur_map = filtered_df[valid_b_mask].groupby(col_D)[col_E].sum() / 60
        
        filtered_df["Frequency"] = filtered_df[col_D].map(freq_map).fillna(0).astype(int)
        filtered_df["Duration_Mins"] = filtered_df[col_D].map(dur_map).fillna(0).round(2)

        summary_common_b_str = ", ".join([format_with_alias(n, alias_map) for n in extracted_common]) if not is_single_file and extracted_common else "None"
        if is_single_file: summary_common_b_str = "N/A (Single File)"

        filtered_df["Has_Multiple_IMEI"] = filtered_df[col_C].apply(lambda x: "Yes" if target_imei_counts.get(x, 0) >= 3 else "No")
        
        display_df = filtered_df.drop_duplicates(subset=[col_D], keep="first").copy()
        display_df['_dt'] = pd.to_datetime(display_df[col_A], errors='coerce')
        display_df = display_df.sort_values(by=['_dt'], ascending=False)
        display_df[col_A] = display_df['_dt'].dt.strftime("%Y-%m-%d %H:%M:%S").fillna("")

        top_b_data = []
        for _, row in display_df[~display_df[col_D].str.startswith('[?]', na=False)].head(10).iterrows():
            top_b_data.append({"b_party": format_with_alias(row[col_D], alias_map), "frequency": str(row["Frequency"]), "last_called": str(row[col_A])})

        area_clusters = [{"area": str(k), "count": int(v)} for k, v in combined_df[col_L].value_counts().head(12).items() if str(k).strip() != '' and str(k).lower() != 'nan']
        
        # Optimize: Convert preview data to a JSON string directly for faster Chaquopy transfer
        preview_list = []
        # Sort by Frequency to show "Top" B-parties
        preview_df = display_df.sort_values(by="Frequency", ascending=False).head(100)
        
        for _, row in preview_df.iterrows():
            bp_val = str(row[col_D]).strip()
            # Filter: Only mobile numbers (starts with 0, length 11)
            if len(bp_val) == 11 and bp_val.startswith('0'):
                preview_list.append({
                    "dt": str(row[col_A]),
                    "ap": format_with_alias(row[col_C], alias_map),
                    "bp": format_with_alias(row[col_D], alias_map),
                    "freq": str(row["Frequency"]),
                    "loc": str(row[col_L])
                })
        preview_data_json = json.dumps(preview_list, ensure_ascii=False)

        hourly_activity = {}
        for a_p in unique_a_parties:
            a_df = combined_df[combined_df[col_C] == a_p].copy()
            a_df['hour'] = pd.to_datetime(a_df[col_A], errors='coerce').dt.hour
            dist = a_df['hour'].value_counts().reindex(range(24), fill_value=0).to_dict()
            hourly_activity[a_p] = {str(h): int(v) for h, v in dist.items()}

        output_filename = f"{''.join(c for c in (unique_a_parties[0] if unique_a_parties else 'Unknown') if c.isalnum())}-CDR-{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_excel_path = os.path.join(output_dir, output_filename)
        with pd.ExcelWriter(output_excel_path, engine="openpyxl") as writer:
            vertical_summary = {
                "Metric": ["A-Parties", "Top Contacts", "Night Stays", "Common Contacts", "IMEI Swaps", "Multi-SIM", "Critical Windows"],
                "Intelligence": [summary_a_parties_str, ", ".join([i["b_party"] for i in top_b_data]), summary_night_stays_str, summary_common_b_str, summary_imei_swappers_str, summary_multi_sim_str, summary_night_routine_str]
            }
            pd.DataFrame(vertical_summary).astype(str).to_excel(writer, sheet_name="Summary", index=False)
            
            excel_df = filtered_df.copy()
            excel_df[col_C] = excel_df[col_C].apply(lambda x: format_with_alias(x, alias_map))
            excel_df[col_D] = excel_df[col_D].apply(lambda x: format_with_alias(x, alias_map))
            excel_df.astype(str).to_excel(writer, sheet_name="Data", index=False)
            for ws in [writer.sheets["Summary"], writer.sheets["Data"]]:
                for row in ws.iter_rows():
                    for cell in row: cell.number_format = "@"

        a_areas = combined_df.groupby(col_C)[col_L].agg(lambda x: x.value_counts().index[0] if not x.empty else "Unknown").to_dict()
        b_areas = combined_df.groupby(col_D)[col_L].agg(lambda x: x.value_counts().index[0] if not x.empty else "Unknown").to_dict()
        
        node_profiles = {}
        for a_p in unique_a_parties:
            a_df = combined_df[combined_df[col_C] == a_p]
            valid_dt = a_df['_parsed_dt'].dropna()
            node_profiles[str(a_p)] = {"total": len(a_df), "first": valid_dt.min().strftime("%Y-%m-%d %H:%M:%S") if not valid_dt.empty else "Unknown", "last": valid_dt.max().strftime("%Y-%m-%d %H:%M:%S") if not valid_dt.empty else "Unknown", "top_loc": str(a_areas.get(a_p, "Unknown")), "alias": alias_map.get(str(a_p), "")}
        
        for b_p in combined_df[col_D].unique():
            sbp = str(b_p)
            if sbp not in node_profiles and sbp != 'nan':
                b_df = combined_df[combined_df[col_D] == b_p]
                valid_dt = b_df['_parsed_dt'].dropna()
                node_profiles[sbp] = {"total": len(b_df), "first": valid_dt.min().strftime("%Y-%m-%d %H:%M:%S") if not valid_dt.empty else "Unknown", "last": valid_dt.max().strftime("%Y-%m-%d %H:%M:%S") if not valid_dt.empty else "Unknown", "top_loc": str(b_areas.get(b_p, "Unknown")), "alias": alias_map.get(sbp, "")}

        sim_to_imei_map = {str(ap): [{"imei": str(i), "hw": lookup_imei(i) or "Generic Handset", "alias": alias_map.get(str(i), "")} for i in imeis] for ap, imeis in ap_to_imeis.items() if str(ap) != 'nan'}

        graph_data = json.dumps({
            "centers": unique_a_parties, 
            "uncommon-links": [{"source": a, "target-links": list(set([str(row[col_D]) for _, row in combined_df[combined_df[col_C] == a].iterrows() if str(row[col_D]) not in extracted_common]))} for a in unique_a_parties],
            "common-links": [{"target": cb, "source": list(set([str(row[col_C]) for _, row in combined_df[combined_df[col_D] == cb].iterrows()]))} for cb in extracted_common],
            "area_clusters": area_clusters, 
            "all_party_areas": {**a_areas, **b_areas},
            "node_profiles": node_profiles,
            "imei_to_sim_map": detailed_imei_map,
            "sim_to_imei_graph": sim_to_imei_data,
            "sim_to_imei_map": sim_to_imei_map,
            "preview_rows_json": preview_data_json,
            "alias_map": alias_map,
            "hourly_activity": hourly_activity
        }, ensure_ascii=False)

        return {"status": "success", "output_path": output_excel_path, "metrics": {"a_parties": summary_a_parties_str, "top_b_parties": top_b_data, "night_stays": summary_night_stays_str, "common_b_parties": summary_common_b_str, "imei_swappers": summary_imei_swappers_str, "multi_sim": summary_multi_sim_str, "night_routine": summary_night_routine_str, "area_clusters": area_clusters, "hourly_activity": hourly_activity, "preview_rows_json": preview_data_json, "graph_data": graph_data}}
    except Exception as e: return {"status": "error", "message": f"Engine failure: {str(e)}"}

def export_same_location_to_excel(json_data, output_path):
    try:
        data = json.loads(json_data)
        if not data: return {"status": "error", "message": "No data to export."}
        df = pd.DataFrame(data)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Same Location Analysis")
            ws = writer.sheets["Same Location Analysis"]
            for row in ws.iter_rows():
                for cell in row: cell.number_format = "@"
        return {"status": "success", "output_path": output_path}
    except Exception as e: return {"status": "error", "message": str(e)}

def search_cdr_data(file_paths, search_query):
    try:
        if not file_paths: return {"status": "error", "message": "No files."}
        alias_map = load_aliases()
        terms = [s.strip() for s in str(search_query).split(",") if s.strip()]
        all_dfs = []
        for path in file_paths:
            if not os.path.exists(path): continue
            try:
                if path.lower().endswith('.pdf'):
                    df = pdf_converter.pdf_to_dataframe(path)
                else:
                    df = pd.read_excel(path, engine="openpyxl", dtype=str)
                
                if df is None or df.empty or len(df.columns) < 12: continue
                nums = df[df.columns[2]].apply(lambda x: "".join(c for c in str(x) if c.isdigit())[-11:]).unique()
                a_party = nums[-1] if len(nums) > 0 else "Unknown"
                df['_internal_a'] = a_party
                df['_internal_time'] = pd.to_datetime(df[df.columns[0]], errors='coerce')
                df['_internal_loc'] = df[df.columns[11]].fillna('').astype(str).str.strip()
                all_dfs.append(df)
            except: continue
        if not all_dfs: return {"status": "error", "message": "No data extracted."}
        combined = pd.concat(all_dfs, ignore_index=True)
        dialog_lines = []
        for term in terms:
            mask = combined.astype(str).apply(lambda col: col.str.contains(term, case=False, na=False)).any(axis=1)
            match_df = combined[mask]
            if not match_df.empty:
                suspects = sorted(list(set(match_df['_internal_a'].astype(str).tolist())))
                hours = match_df['_internal_time'].dt.hour
                night_count = len(hours[(hours >= 18) | (hours < 6)])
                intensity = "Night Heavy" if night_count > len(match_df)/2 else "Day Heavy"
                locs = match_df['_internal_loc'][match_df['_internal_loc'] != ''].value_counts().head(3)
                loc_str = ", ".join(locs.index) if not locs.empty else "N/A"
                dialog_lines.append(f"<b>Term: <font color='#3182CE'>{term}</font></b><br/>• Linked Suspects: {', '.join([format_with_alias(s, alias_map) for s in suspects])}<br/>• Frequency: {len(match_df)} hits ({intensity})<br/>• Top Locations: {loc_str}")
            else: dialog_lines.append(f"<font color='#E53E3E'><b>Term: {term}</b></font><br/>&nbsp;&nbsp;Status: <i>Not Found</i>")
        return {"status": "success", "summary_html": "<br/><br/>".join(dialog_lines)}
    except Exception as e: return {"status": "error", "message": str(e)}

def crop_cdr_data(file_paths, location_query, start_ts, end_ts, output_dir):
    try:
        all_dfs = []
        for p in file_paths:
            if not os.path.exists(p): continue
            try:
                if p.lower().endswith('.pdf'):
                    df = pdf_converter.pdf_to_dataframe(p)
                else:
                    df = pd.read_excel(p, engine="openpyxl", dtype=str)
                if not df.empty: all_dfs.append(df)
            except: continue
            
        if not all_dfs: return {"status": "error", "message": "No data."}
        combined = pd.concat(all_dfs, ignore_index=True)
        temp_time = pd.to_datetime(combined[combined.columns[0]], errors='coerce')
        mask = (temp_time >= pd.to_datetime(start_ts)) & (temp_time <= pd.to_datetime(end_ts))
        if location_query.strip(): mask &= combined[combined.columns[11]].fillna('').astype(str).str.contains(location_query, case=False, na=False)
        cropped = combined[mask].copy().astype(str).replace('nan', '')
        if cropped.empty: return {"status": "error", "message": "No matches."}
        out_path = os.path.join(output_dir, f"{time.strftime('%Y%m%d_%H%M%S')}_{len(file_paths)}CDRsCropped.xlsx")
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            cropped.to_excel(writer, index=False)
            for row in writer.sheets['Sheet1'].iter_rows():
                for cell in row: cell.number_format = '@'
        return {"status": "success", "output_path": out_path, "count": len(cropped)}
    except Exception as e: return {"status": "error", "message": str(e)}

def same_location_analysis(file_paths, progress_callback=None):
    try:
        paths = [str(p).strip() for p in file_paths if p and str(p) != 'None' and len(str(p)) > 5]
        alias_map = load_aliases()
        if len(paths) < 2: return {"status": "error", "message": f"Need at least 2 CDRs. Found: {len(paths)}."}
        all_data = []
        for p in paths:
            if os.path.exists(p):
                try:
                    if p.lower().endswith('.pdf'):
                        df = pdf_converter.pdf_to_dataframe(p)
                    else:
                        df = pd.read_excel(p, engine="openpyxl", dtype=str)
                except:
                    continue

                if not df.empty and len(df.columns) >= 12:
                    t = pd.DataFrame()
                    # Try flexible date parsing for PDF/Excel mixed sources
                    t['R'] = pd.to_datetime(df[df.columns[0]], errors='coerce', dayfirst=True)
                    # If parsing failed for some rows, try without dayfirst
                    mask = t['R'].isna()
                    if mask.any():
                        t.loc[mask, 'R'] = pd.to_datetime(df.loc[mask, df.columns[0]], errors='coerce')
                    
                    t['D'] = t['R'].dt.date
                    t['S'] = df[df.columns[0]].fillna('').astype(str)
                    t['A'] = df[df.columns[2]].fillna('').astype(str).str.strip().apply(lambda x: "".join(c for c in str(x) if c.isdigit())[-11:])
                    t['B'] = df[df.columns[3]].fillna('').astype(str).str.strip().apply(lambda x: "".join(c for c in str(x) if c.isdigit())[-11:])
                    t['L'] = df[df.columns[7]].fillna('').astype(str).str.strip().apply(lambda x: str(x).split('.')[0])
                    t['C'] = df[df.columns[8]].fillna('').astype(str).str.strip().apply(lambda x: str(x).split('.')[0])
                    t['Loc'] = df[df.columns[11]].fillna('').astype(str).str.strip()
                    
                    # Basic Validation: Must have a valid Date and A Party
                    t = t[t['D'].notna() & (t['A'] != '')]
                    if not t.empty:
                        all_data.append(t)
        if len(all_data) < 2: return {"status": "error", "message": "Insufficient valid data."}
        combined = pd.concat(all_data, ignore_index=True).dropna(subset=['D', 'A'])
        results = []
        unique_days = sorted(combined['D'].unique())
        total_days = len(unique_days)
        for i, d in enumerate(unique_days):
            day_df = combined[combined['D'] == d]
            if day_df['A'].nunique() < 2: 
                if progress_callback: progress_callback.onProgress(int((i + 1) / total_days * 100))
                continue
            valid_lac = day_df[(day_df['L'] != '') & (day_df['L'] != 'nan')]
            if not valid_lac.empty:
                matches = valid_lac.groupby('L').filter(lambda x: x['A'].nunique() > 1)
                for _, row in matches.iterrows():
                    is_c = day_df[(day_df['L'] == row['L']) & (day_df['C'] == row['C'])]['A'].nunique() > 1
                    results.append({"Time": row['S'], "A_Party": row['A'], "B_Party": row['B'], "LAC": row['L'], "Cell": row['C'], "BTS_Loc": row['Loc'], "Reason": "LAC+Cell Match" if is_c else "LAC Match"})
            a_list = sorted(list(day_df['A'].unique()))
            addr_map = day_df[day_df['Loc'] != ''].groupby('A')['Loc'].unique().to_dict()
            for i_ap in range(len(a_list)):
                for j_ap in range(i_ap+1, len(a_list)):
                    ap1, ap2 = a_list[i_ap], a_list[j_ap]
                    for ad1 in addr_map.get(ap1, []):
                        for ad2 in addr_map.get(ap2, []):
                            if SequenceMatcher(None, str(ad1).lower(), str(ad2).lower()).ratio() >= 0.7:
                                rows = day_df[(day_df['A'].isin([ap1, ap2])) & (day_df['Loc'].isin([ad1, ad2]))]
                                for _, r in rows.iterrows():
                                    results.append({"Time": r['S'], "A_Party": r['A'], "B_Party": r['B'], "LAC": r['L'], "Cell": r['C'], "BTS_Loc": r['Loc'], "Reason": "Tower Similarity (>70%)"})
            if progress_callback:
                progress_callback.onProgress(int((i + 1) / total_days * 100))

        if not results: return {"status": "success", "data": "[]", "summary": "No spatial overlaps or concurrent tower hits detected between these parties."}
        final_df = pd.DataFrame(results).drop_duplicates(subset=['Time', 'A_Party'])
        try:
            summary_df = final_df.copy()
            summary_df['_dt'] = pd.to_datetime(summary_df['Time'], errors='coerce')
            summary_df = summary_df.dropna(subset=['_dt'])
            summary_df['_bucket'] = summary_df['_dt'].dt.floor('30T')
            groups = summary_df.groupby(['_bucket', 'LAC', 'Cell', 'BTS_Loc'])['A_Party'].unique()
            set_counts = {}
            for idx, parties in groups.items():
                if len(parties) < 2: continue
                p_list = sorted([format_with_alias(p, alias_map) for p in parties])
                p_key = " and ".join(p_list) if len(p_list) == 2 else ", ".join(p_list[:-1]) + " and " + p_list[-1]
                if p_key not in set_counts: set_counts[p_key] = set()
                set_counts[p_key].add(idx[0].strftime("%d/%m/%y %H:%M"))
            summary_lines = []
            import string
            sorted_sets = sorted(set_counts.items(), key=lambda x: len(x[1]), reverse=True)
            for i, (p_key, times_set) in enumerate(sorted_sets):
                label = list(string.ascii_lowercase)[i % 26] if i < 26 else f"z{i}"
                times_list = sorted(list(times_set), reverse=True)
                summary_lines.append(f"{label}. Parties {p_key} were found around each other [{len(times_list)}] times. The times are: {' | '.join(times_list[:5])}")
            spatial_summary = "\n".join(summary_lines) if summary_lines else "No concurrent overlaps."
        except Exception as e: spatial_summary = f"Summary error: {str(e)}"
        
        data_list = final_df.sort_values('Time', ascending=False).head(1500).to_dict('records')
        for item in data_list:
            item['A_Party'] = format_with_alias(item['A_Party'], alias_map)
            item['B_Party'] = format_with_alias(item['B_Party'], alias_map)
        return {"status": "success", "data": json.dumps(data_list, ensure_ascii=False), "summary": spatial_summary}
    except Exception as e: return {"status": "error", "message": f"Critical Error: {str(e)}"}
